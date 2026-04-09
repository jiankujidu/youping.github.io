# -*- coding: utf-8 -*-
"""
ELISA代测表填写系统 - 后端服务
复制原始Excel模板并填充用户数据，保留所有样式
"""

import os
import sys

# 设置控制台编码
if sys.platform == 'win32':
    import locale
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

from http.server import HTTPServer, BaseHTTPRequestHandler
import json
from openpyxl import load_workbook
from io import BytesIO
import base64

# 原始Excel模板路径（使用相对路径，与server.py同目录）
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(SCRIPT_DIR, '优品Elisa代测表.xlsx')

# 检查模板文件是否存在
if not os.path.exists(TEMPLATE_PATH):
    print('='*50)
    print('错误：找不到Excel模板文件！')
    print(f'预期路径：{TEMPLATE_PATH}')
    print(f'脚本目录：{SCRIPT_DIR}')
    print(f'目录文件：{os.listdir(SCRIPT_DIR)}')
    print('='*50)
    sys.exit(1)

print(f'模板文件已找到：{TEMPLATE_PATH}')


class ELISAHandler(BaseHTTPRequestHandler):
    def log_message(self, format, *args):
        """自定义日志"""
        pass
    
    def do_OPTIONS(self):
        """处理CORS预检请求"""
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()
    
    def do_POST(self):
        """处理POST请求"""
        if self.path == '/export':
            try:
                # 读取请求体
                content_length = int(self.headers.get('Content-Length', 0))
                body = self.rfile.read(content_length)
                data = json.loads(body.decode('utf-8'))
                
                print(f'收到导出请求，客户：{data.get("name", "未知")}')
                
                # 填充Excel
                excel_data = fill_excel(data)
                
                # 返回base64编码的Excel文件
                self.send_response(200)
                self.send_header('Content-Type', 'application/json; charset=utf-8')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                
                response = json.dumps({
                    'success': True,
                    'data': base64.b64encode(excel_data).decode('utf-8'),
                    'filename': '优品Elisa代测表_已填写.xlsx'
                }, ensure_ascii=False)
                self.wfile.write(response.encode('utf-8'))
                print('导出成功！')
                
            except FileNotFoundError as e:
                print(f'文件错误：{e}')
                self.send_error_response(f'找不到Excel模板文件')
            except Exception as e:
                print(f'导出错误：{e}')
                import traceback
                traceback.print_exc()
                self.send_error_response(str(e))
        else:
            self.send_error(404)
    
    def send_error_response(self, error_msg):
        """发送错误响应"""
        try:
            self.send_response(500)
            self.send_header('Content-Type', 'application/json; charset=utf-8')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            response = json.dumps({'success': False, 'error': error_msg}, ensure_ascii=False)
            self.wfile.write(response.encode('utf-8'))
        except:
            pass
    
    def do_GET(self):
        """处理GET请求"""
        if self.path == '/health':
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(b'{"status": "ok"}')
        else:
            self.send_error(404)


def fill_excel(data):
    """
    填充Excel数据，保留原始样式
    """
    # 加载原始模板
    wb = load_workbook(TEMPLATE_PATH)
    ws1 = wb['样本检测要求信息表']
    
    # ===== 第3行：客户基本信息 =====
    ws1['B3'] = data.get('name', '')
    ws1['D3'] = data.get('phone', '')
    ws1['G3'] = data.get('company', '')
    
    # ===== 第4行：快递和订单信息 =====
    ws1['B4'] = data.get('trackingNo', '')
    ws1['D4'] = data.get('orderNo', '')
    ws1['G4'] = data.get('salesperson', '陈美龙')
    
    # ===== 第8行：种属和样本数量 =====
    ws1['B8'] = data.get('species', '')
    sample_count = data.get('sampleCount', '')
    ws1['G8'] = int(sample_count) if sample_count else ''
    
    # ===== 第9行：样本类型 =====
    ws1['B9'] = data.get('sampleType', '')
    
    # ===== 第10行：指标名称 =====
    ws1['B10'] = data.get('indicator', '')
    
    # ===== 第11行：实验要求和实验目的 =====
    test_req = data.get('testRequirement', '')
    test_purpose = data.get('testPurpose', '')
    ws1['B11'] = f'{test_req} / {test_purpose}'
    
    # ===== 第12行：备注 =====
    ws1['B12'] = data.get('remarks', '')
    
    # ===== 第13行：实验样本信息 =====
    ws1['B13'] = data.get('sampleInfo', '')
    
    # ===== 第14行：样本是否做重复 =====
    ws1['B14'] = data.get('sampleRepeat', '无')
    
    # ===== 第16行：标曲是否做重复 =====
    ws1['B16'] = data.get('standardRepeat', '无')
    
    # ===== Sheet 2: 样本检测信息采集表 =====
    ws2 = wb['样本检测信息采集表']
    
    samples = data.get('samples', [])
    for i, sample in enumerate(samples):
        row = i + 3
        if row <= 82:
            ws2.cell(row=row, column=1, value=i + 1)
            ws2.cell(row=row, column=2, value=sample.get('stype', ''))
            ws2.cell(row=row, column=3, value=sample.get('scode', ''))
            ws2.cell(row=row, column=4, value=sample.get('sgroup', ''))
            ws2.cell(row=row, column=5, value=sample.get('samount', ''))
            ws2.cell(row=row, column=6, value=sample.get('stemp', ''))
            ws2.cell(row=row, column=7, value=sample.get('sdate', ''))
            ws2.cell(row=row, column=8, value=sample.get('snote', ''))
            ws2.cell(row=row, column=9, value=sample.get('sreq', ''))
    
    # 保存到BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def run_server(port=8765):
    """启动HTTP服务器"""
    server_address = ('127.0.0.1', port)
    httpd = HTTPServer(server_address, ELISAHandler)
    
    print('='*50)
    print(f'ELISA代测表服务已启动')
    print(f'服务地址: http://127.0.0.1:{port}')
    print('请在浏览器中打开 index.html 填写表单')
    print('按 Ctrl+C 停止服务')
    print('='*50)
    
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        print('\n服务已停止')
        httpd.shutdown()


if __name__ == '__main__':
    run_server()
